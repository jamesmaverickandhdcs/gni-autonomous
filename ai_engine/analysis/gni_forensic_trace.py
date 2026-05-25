# ============================================================
# GNI Article Forensic Trace Generator
# Produces XLSX showing every article's journey through the pipeline
# One row per article, one column per decision gate
# Sends to admin Telegram as file
#
# Columns:
#   Identity:    Serial, REF-ID, Source, Title, URL, Published, Pillar
#   Stage 1:     Relevance Result, Matched Keywords
#   Stage 1b-A:  Content Type, Type Signals
#   Stage 1b-B:  Injection Result, Injection Reason
#   Stage 1b-C:  Sanitized
#   Stage 1b-D:  Review Gate Result, Gate Reason
#   Stage 2:     Dedup Result, Dedup Reason
#   Stage 3:     Significance Score
#   Stage 4:     Selected, Rank
#   Final:       Final Status (color coded)
#
# Color coding:
#   Green:  SELECTED (made it to AI analysis)
#   Yellow: Passed all gates, not selected (score too low)
#   Orange: Rejected at Stage 1b
#   Red:    Rejected at Stage 1 (irrelevant)
#   Gray:   Not evaluated (failed earlier stage)
# ============================================================

import os
import tempfile
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        GradientFill
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ============================================================
# Color palette
# ============================================================
C_HEADER_BG    = '1B3A6B'   # dark blue header
C_HEADER_FG    = 'FFFFFF'   # white text
C_SELECTED     = 'C6EFCE'   # green - selected
C_SELECTED_FG  = '276221'
C_PASSED       = 'FFFF99'   # yellow - passed not selected
C_PASSED_FG    = '7D6608'
C_REJ_1B       = 'FCE4D6'   # orange - rejected at 1b
C_REJ_1B_FG    = 'C55A11'
C_REJ_1        = 'FFCCCC'   # red - rejected at stage 1
C_REJ_1_FG     = '9C0006'
C_GRAY         = 'F2F2F2'   # gray - not evaluated
C_GRAY_FG      = '7F7F7F'
C_PASS_CELL    = 'E2EFDA'   # light green cell
C_FAIL_CELL    = 'FFCCCC'   # light red cell
C_SKIP_CELL    = 'F2F2F2'   # gray cell
C_SCORE_HIGH   = 'C6EFCE'   # green score
C_SCORE_MED    = 'FFEB9C'   # yellow score
C_SCORE_LOW    = 'FFCCCC'   # red score
C_ALT_ROW      = 'F9F9F9'   # alternating row
C_WHITE        = 'FFFFFF'
C_BLUE_LIGHT   = 'DEEAF1'

C_GEO          = 'D5E8F0'
C_TECH         = 'E8F5E9'
C_FIN          = 'FFF3CD'
C_OTHER        = 'F2F2F2'


def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color='000000', size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Arial')


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _pillar_color(pillar):
    p = (pillar or '').lower()
    if p == 'geo': return C_GEO
    if p == 'tech': return C_TECH
    if p == 'fin': return C_FIN
    return C_OTHER


def _final_status(art):
    """Determine final status label and color for each article."""
    if art.get('stage4_selected'):
        return 'SELECTED', C_SELECTED, C_SELECTED_FG
    if not art.get('stage1_passed', False):
        return 'REJECTED @ STAGE 1', C_REJ_1, C_REJ_1_FG
    if not art.get('stage1b_passed', True):
        return 'REJECTED @ STAGE 1b', C_REJ_1B, C_REJ_1B_FG
    if not art.get('stage2_passed', True):
        return 'REJECTED @ STAGE 2', C_REJ_1B, C_REJ_1B_FG
    # Passed all gates but not selected
    return 'PASSED — NOT SELECTED', C_PASSED, C_PASSED_FG


def _row_bg(art):
    status, color, _ = _final_status(art)
    return color


def _cell_result(passed, skip=False):
    """Returns (text, bg_color) for a pass/fail/skip cell."""
    if skip:
        return 'SKIP', C_SKIP_CELL
    if passed:
        return 'PASS', C_PASS_CELL
    return 'FAIL', C_FAIL_CELL


def _score_color(score):
    if score is None: return C_SKIP_CELL
    s = float(score)
    if s >= 20: return C_SCORE_HIGH
    if s >= 10: return C_SCORE_MED
    return C_SCORE_LOW


def generate_forensic_trace(articles: list, run_id: str,
                             run_at: str, pipeline_meta: dict) -> str | None:
    """
    Generate the Article Forensic Trace XLSX.
    articles: list of dicts from pipeline_articles Supabase query
    Returns path to generated file, or None on failure.
    """
    if not OPENPYXL_OK:
        print('  Warning: openpyxl not installed -- forensic trace skipped')
        return None

    if not articles:
        print('  Warning: No articles to trace')
        return None

    try:
        wb = Workbook()

        # ============================================================
        # SHEET 1: SELECTED ARTICLES (Quick View)
        # ============================================================
        ws_sel = wb.active
        ws_sel.title = 'Selected Articles'
        _build_selected_sheet(ws_sel, articles, run_at, pipeline_meta)

        # ============================================================
        # SHEET 2: FULL FORENSIC TRACE
        # ============================================================
        ws_trace = wb.create_sheet('Full Forensic Trace')
        _build_trace_sheet(ws_trace, articles, run_at, pipeline_meta)

        # ============================================================
        # SHEET 3: REJECTION ANALYSIS
        # ============================================================
        ws_rej = wb.create_sheet('Rejection Analysis')
        _build_rejection_sheet(ws_rej, articles, run_at)

        # ============================================================
        # SHEET 4: SOURCE PERFORMANCE
        # ============================================================
        ws_src = wb.create_sheet('Source Performance')
        _build_source_sheet(ws_src, articles, run_at)

        # Save
        date_str = run_at[:10]
        time_str = run_at[11:16].replace(':', '')
        filename = f'{date_str}_GNI_Forensic_Trace_{time_str}_UTC.xlsx'
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)

        size_kb = os.path.getsize(output_path) // 1024
        print(f'  OK Forensic trace generated: {filename} ({size_kb}KB)')
        return output_path

    except Exception as e:
        print(f'  Warning: Forensic trace generation error: {str(e)[:80]}')
        import traceback
        traceback.print_exc()
        return None


def _build_selected_sheet(ws, articles, run_at, pipeline_meta):
    """Sheet 1: Quick view of selected articles only."""
    selected = [a for a in articles if a.get('stage4_selected')]
    selected.sort(key=lambda x: x.get('stage4_rank') or 99)

    date_str = run_at[:10]
    time_str = run_at[11:16] + ' UTC'

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f'GNI — Selected Articles  |  {date_str}  {time_str}  |  {len(selected)} articles selected from {len(articles)} collected'
    ws['A1'].font = _font(bold=True, color=C_HEADER_FG, size=12)
    ws['A1'].fill = _fill(C_HEADER_BG)
    ws['A1'].alignment = _align('center')
    ws.row_dimensions[1].height = 22

    # Pipeline summary row
    ws.merge_cells('A2:I2')
    collected = pipeline_meta.get('articles_collected', len(articles))
    after_rel = pipeline_meta.get('after_relevance', '?')
    after_1b = pipeline_meta.get('after_1b', '?')
    after_dedup = pipeline_meta.get('after_dedup', '?')
    ws['A2'] = f'Pipeline: {collected} collected → relevance filter → {after_1b} after Stage 1b → {after_dedup} after dedup → {len(selected)} selected'
    ws['A2'].font = _font(italic=True, color='444444', size=9)
    ws['A2'].fill = _fill(C_BLUE_LIGHT)
    ws['A2'].alignment = _align('center')

    # Headers
    headers = ['Rank', 'REF-ID', 'Source', 'Pillar', 'Title',
               'Score', 'Content Type', 'Final Status', 'URL']
    col_widths = [6, 18, 18, 8, 60, 8, 16, 20, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _align('center')
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[3].height = 18
    ws.freeze_panes = 'A4'

    # Data rows
    for row_idx, art in enumerate(selected, 4):
        rank = art.get('stage4_rank') or row_idx - 3
        ref_id = _make_ref_id(art, run_at)
        source = art.get('source', '')
        pillar = (art.get('pillar') or '').upper()
        title = art.get('title', '')
        score = art.get('stage3_score', 0)
        ct = art.get('content_type', 'news')
        url = art.get('url', '')

        row_data = [rank, ref_id, source, pillar, title,
                    score, ct, 'SELECTED', url]

        bg = _pillar_color(art.get('pillar', ''))

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _fill(bg)
            cell.font = _font(size=9)
            cell.alignment = _align('left', wrap=(col == 5))
            cell.border = _border_thin()

        # Score color
        score_cell = ws.cell(row=row_idx, column=6)
        score_cell.fill = _fill(_score_color(score))
        score_cell.font = _font(bold=True, size=9)
        score_cell.alignment = _align('center')

        # Status cell always green
        ws.cell(row=row_idx, column=8).fill = _fill(C_SELECTED)
        ws.cell(row=row_idx, column=8).font = _font(bold=True, color=C_SELECTED_FG, size=9)

        ws.row_dimensions[row_idx].height = 30

    # Auto-filter
    ws.auto_filter.ref = f'A3:I{3 + len(selected)}'


def _build_trace_sheet(ws, articles, run_at, pipeline_meta):
    """Sheet 2: Full forensic trace — every article, every gate."""

    date_str = run_at[:10]
    time_str = run_at[11:16] + ' UTC'

    # Title
    ws.merge_cells('A1:T1')
    ws['A1'] = f'GNI Article Forensic Trace  |  {date_str} {time_str}  |  {len(articles)} articles documented'
    ws['A1'].font = _font(bold=True, color=C_HEADER_FG, size=12)
    ws['A1'].fill = _fill(C_HEADER_BG)
    ws['A1'].alignment = _align('center')
    ws.row_dimensions[1].height = 22

    # Column groups and headers
    col_groups = [
        # (header, width, group_label)
        ('Serial',          6,  'IDENTITY'),
        ('REF-ID',          18, 'IDENTITY'),
        ('Source',          18, 'IDENTITY'),
        ('Pillar',          7,  'IDENTITY'),
        ('Title',           50, 'IDENTITY'),
        ('Published',       14, 'IDENTITY'),
        ('URL',              40, 'IDENTITY'),
        ('S1 Result',       10, 'STAGE 1\nRelevance'),
        ('S1 Keywords',     25, 'STAGE 1\nRelevance'),
        ('Content Type',    14, 'STAGE 1b-A\nClassifier'),
        ('Type Signals',    25, 'STAGE 1b-A\nClassifier'),
        ('Injection',       10, 'STAGE 1b-B\nInjection'),
        ('Inject Reason',   25, 'STAGE 1b-B\nInjection'),
        ('Sanitized',       9,  'STAGE 1b-C\nSanitize'),
        ('Review Gate',     12, 'STAGE 1b-D\nReview Gate'),
        ('Gate Reason',     30, 'STAGE 1b-D\nReview Gate'),
        ('Dedup',           10, 'STAGE 2\nDedup'),
        ('Dedup Reason',    20, 'STAGE 2\nDedup'),
        ('Score',           8,  'STAGE 3\nScoring'),
        ('Rank',            6,  'STAGE 4\nSelection'),
        ('FINAL STATUS',    22, 'FINAL'),
    ]

    # Group header row (row 2)
    group_colors = {
        'IDENTITY':          '4472C4',
        'STAGE 1\nRelevance': '70AD47',
        'STAGE 1b-A\nClassifier': 'ED7D31',
        'STAGE 1b-B\nInjection': 'FF0000',
        'STAGE 1b-C\nSanitize':  'FFC000',
        'STAGE 1b-D\nReview Gate': 'C55A11',
        'STAGE 2\nDedup':    '7030A0',
        'STAGE 3\nScoring':  '2E75B6',
        'STAGE 4\nSelection': '375623',
        'FINAL':             '1B3A6B',
    }

    last_group = None
    group_start_col = 1
    for col, (_, width, group) in enumerate(col_groups, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
        if group != last_group:
            # Write group header
            ws.cell(row=2, column=col, value=group.replace('\n', ' — '))
            cell = ws.cell(row=2, column=col)
            cell.font = _font(bold=True, color='FFFFFF', size=9)
            cell.fill = _fill(group_colors.get(group, '444444'))
            cell.alignment = _align('center')
            cell.border = _border_thin()
            last_group = group
        else:
            cell = ws.cell(row=2, column=col)
            cell.fill = _fill(group_colors.get(group, '444444'))
            cell.border = _border_thin()

    ws.row_dimensions[2].height = 16

    # Column headers (row 3)
    for col, (header, _, group) in enumerate(col_groups, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = _font(bold=True, color='FFFFFF', size=9)
        cell.fill = _fill(group_colors.get(group, '444444'))
        cell.alignment = _align('center', wrap=True)
        cell.border = _border_thin()

    ws.row_dimensions[3].height = 22
    ws.freeze_panes = 'A4'

    # Sort: selected first, then by score desc
    def sort_key(a):
        sel = 0 if a.get('stage4_selected') else 1
        score = -(a.get('stage3_score') or 0)
        return (sel, score)

    sorted_arts = sorted(articles, key=sort_key)

    # Data rows
    for row_idx, art in enumerate(sorted_arts, 4):
        serial = row_idx - 3
        ref_id = _make_ref_id(art, run_at)
        status_label, row_color, fg_color = _final_status(art)

        # Parse stage1b_reason for injection/sanitize/gate info
        s1b_reason = art.get('stage1b_reason', '') or ''
        injection_result, injection_reason = _parse_injection(s1b_reason, art)
        sanitized = _parse_sanitized(s1b_reason)
        gate_result, gate_reason = _parse_gate(art)

        row_vals = [
            serial,                                    # 1 Serial
            ref_id,                                    # 2 REF-ID
            art.get('source', ''),                     # 3 Source
            (art.get('pillar') or '').upper(),         # 4 Pillar
            art.get('title', ''),                      # 5 Title
            art.get('url', ''),                        # URL
            str(art.get('published_at', ''))[:16],     # 6 Published
            'PASS' if art.get('stage1_passed') else 'FAIL',  # 7 S1 Result
            art.get('stage1_reason', ''),              # 8 S1 Keywords
            art.get('content_type', 'news'),           # 9 Content Type
            art.get('content_type_signals', ''),       # 10 Type Signals
            injection_result,                          # 11 Injection
            injection_reason,                          # 12 Inject Reason
            'YES' if sanitized else 'NO',              # 13 Sanitized
            gate_result,                               # 14 Review Gate
            gate_reason,                               # 15 Gate Reason
            'PASS' if art.get('stage2_passed', True) else 'FAIL',  # 16 Dedup
            art.get('stage2_reason', '') or '',        # 17 Dedup Reason
            art.get('stage3_score', 0),                # 18 Score
            art.get('stage4_rank') or '',              # 19 Rank
            status_label,                              # 20 Final Status
        ]

        # Determine alternating row base color
        alt = C_WHITE if serial % 2 == 0 else C_ALT_ROW

        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = _font(size=8)
            cell.alignment = _align('left', wrap=(col in [5, 9, 11, 13, 16, 18]))
            cell.border = _border_thin()

            # Default background
            cell.fill = _fill(alt)

        # Column-specific formatting
        # S1 Result (col 7)
        s1_passed = art.get('stage1_passed', False)
        ws.cell(row=row_idx, column=7).fill = _fill(C_PASS_CELL if s1_passed else C_FAIL_CELL)
        ws.cell(row=row_idx, column=7).font = _font(bold=True, size=8,
            color=C_SELECTED_FG if s1_passed else '9C0006')
        ws.cell(row=row_idx, column=7).alignment = _align('center')

        # Content type (col 9)
        ct = art.get('content_type', 'news')
        ct_colors = {'news': C_PASS_CELL, 'news_with_review': 'FFEB9C', 'review_only': C_REJ_1B}
        ws.cell(row=row_idx, column=9).fill = _fill(ct_colors.get(ct, C_SKIP_CELL))
        ws.cell(row=row_idx, column=9).alignment = _align('center')

        # Injection (col 11)
        ws.cell(row=row_idx, column=11).fill = _fill(
            C_PASS_CELL if injection_result == 'PASS' else
            C_FAIL_CELL if injection_result == 'FAIL' else C_SKIP_CELL)
        ws.cell(row=row_idx, column=11).alignment = _align('center')

        # Sanitized (col 13)
        ws.cell(row=row_idx, column=13).fill = _fill('FFEB9C' if sanitized else C_PASS_CELL)
        ws.cell(row=row_idx, column=13).alignment = _align('center')

        # Review Gate (col 14)
        gt_colors = {'PASS': C_PASS_CELL, 'FAIL': C_FAIL_CELL,
                     'SKIP': C_SKIP_CELL, 'REJECTED': C_FAIL_CELL}
        ws.cell(row=row_idx, column=14).fill = _fill(gt_colors.get(gate_result, C_SKIP_CELL))
        ws.cell(row=row_idx, column=14).alignment = _align('center')

        # Dedup (col 16)
        d_passed = art.get('stage2_passed', True)
        ws.cell(row=row_idx, column=16).fill = _fill(C_PASS_CELL if d_passed else C_FAIL_CELL)
        ws.cell(row=row_idx, column=16).alignment = _align('center')

        # Score (col 18)
        score = art.get('stage3_score', 0)
        ws.cell(row=row_idx, column=18).fill = _fill(_score_color(score))
        ws.cell(row=row_idx, column=18).font = _font(bold=True, size=8)
        ws.cell(row=row_idx, column=18).alignment = _align('center')

        # Final status (col 20)
        ws.cell(row=row_idx, column=20).fill = _fill(row_color)
        ws.cell(row=row_idx, column=20).font = _font(bold=True, color=fg_color, size=8)
        ws.cell(row=row_idx, column=20).alignment = _align('center')

        ws.row_dimensions[row_idx].height = 28

    # Auto-filter on header row
    ws.auto_filter.ref = f'A3:U{3 + len(articles)}'


def _build_rejection_sheet(ws, articles, run_at):
    """Sheet 3: Rejection analysis — why articles were rejected."""
    rejected = [a for a in articles if not a.get('stage4_selected')]

    ws.merge_cells('A1:F1')
    ws['A1'] = f'Rejection Analysis  |  {run_at[:10]}  |  {len(rejected)} articles rejected'
    ws['A1'].font = _font(bold=True, color=C_HEADER_FG, size=11)
    ws['A1'].fill = _fill(C_HEADER_BG)
    ws['A1'].alignment = _align('center')

    # Summary table
    stages = {
        'Stage 1 — Irrelevant': len([a for a in articles if not a.get('stage1_passed')]),
        'Stage 1b — Injection': len([a for a in articles
                                      if a.get('stage1_passed') and not a.get('stage1b_passed')
                                      and 'injection' in (a.get('stage1b_reason') or '').lower()]),
        'Stage 1b — Review Gate': len([a for a in articles
                                        if a.get('stage1_passed') and not a.get('stage1b_passed')
                                        and 'gate' in (a.get('stage1b_reason') or '').lower()]),
        'Stage 2 — Duplicate': len([a for a in articles
                                     if a.get('stage1_passed') and a.get('stage1b_passed')
                                     and not a.get('stage2_passed', True)]),
        'Stage 3/4 — Score Too Low': len([a for a in articles
                                           if a.get('stage1_passed') and a.get('stage1b_passed')
                                           and a.get('stage2_passed', True)
                                           and not a.get('stage4_selected')]),
    }

    ws['A2'] = 'Rejection Stage'
    ws['B2'] = 'Count'
    ws['C2'] = '% of Total'
    for col in range(1, 4):
        cell = ws.cell(row=2, column=col)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=10)
        cell.fill = _fill('2E75B6')
        cell.alignment = _align('center')
        cell.border = _border_thin()

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12

    total = len(articles)
    stage_colors = {
        'Stage 1 — Irrelevant':       C_REJ_1,
        'Stage 1b — Injection':       C_FAIL_CELL,
        'Stage 1b — Review Gate':     C_REJ_1B,
        'Stage 2 — Duplicate':        'E8D5F5',
        'Stage 3/4 — Score Too Low':  C_PASSED,
    }

    for row_idx, (stage, count) in enumerate(stages.items(), 3):
        ws.cell(row=row_idx, column=1, value=stage).fill = _fill(stage_colors.get(stage, C_WHITE))
        ws.cell(row=row_idx, column=2, value=count).fill = _fill(stage_colors.get(stage, C_WHITE))
        pct = f'{count/total*100:.1f}%' if total > 0 else '0%'
        ws.cell(row=row_idx, column=3, value=pct).fill = _fill(stage_colors.get(stage, C_WHITE))
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).font = _font(size=9)
            ws.cell(row=row_idx, column=col).alignment = _align('center')
            ws.cell(row=row_idx, column=col).border = _border_thin()


def _build_source_sheet(ws, articles, run_at):
    """Sheet 4: Source performance — how each source is doing."""
    from collections import defaultdict

    source_stats = defaultdict(lambda: {
        'collected': 0, 'passed_s1': 0, 'passed_s1b': 0,
        'passed_s2': 0, 'selected': 0, 'total_score': 0.0
    })

    for art in articles:
        src = art.get('source', 'Unknown')
        source_stats[src]['collected'] += 1
        if art.get('stage1_passed'):
            source_stats[src]['passed_s1'] += 1
        if art.get('stage1_passed') and art.get('stage1b_passed', True):
            source_stats[src]['passed_s1b'] += 1
        if art.get('stage1_passed') and art.get('stage1b_passed', True) and art.get('stage2_passed', True):
            source_stats[src]['passed_s2'] += 1
        if art.get('stage4_selected'):
            source_stats[src]['selected'] += 1
            source_stats[src]['total_score'] += float(art.get('stage3_score') or 0)

    ws.merge_cells('A1:H1')
    ws['A1'] = f'Source Performance  |  {run_at[:10]}  |  {len(source_stats)} sources'
    ws['A1'].font = _font(bold=True, color=C_HEADER_FG, size=11)
    ws['A1'].fill = _fill(C_HEADER_BG)
    ws['A1'].alignment = _align('center')

    headers = ['Source', 'Collected', 'Pass S1', 'Pass S1b', 'Pass S2',
               'Selected', 'Avg Score', 'Geo Ratio']
    col_widths = [25, 10, 10, 10, 10, 10, 10, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=9)
        cell.fill = _fill('2E75B6')
        cell.alignment = _align('center')
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A3'

    # Sort by selected desc, then collected desc
    sorted_sources = sorted(source_stats.items(),
                             key=lambda x: (-x[1]['selected'], -x[1]['collected']))

    for row_idx, (src, stats) in enumerate(sorted_sources, 3):
        collected = stats['collected']
        selected = stats['selected']
        avg_score = round(stats['total_score'] / selected, 1) if selected > 0 else 0
        geo_ratio = f"{stats['passed_s1']/collected*100:.0f}%" if collected > 0 else '0%'

        row_data = [src, collected, stats['passed_s1'], stats['passed_s1b'],
                    stats['passed_s2'], selected, avg_score, geo_ratio]

        bg = C_SELECTED if selected > 0 else (C_WHITE if row_idx % 2 == 0 else C_ALT_ROW)

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = _font(size=9, bold=(col == 1))
            cell.fill = _fill(bg)
            cell.alignment = _align('left' if col == 1 else 'center')
            cell.border = _border_thin()

        ws.row_dimensions[row_idx].height = 16


# ============================================================
# Helper parsers
# ============================================================
def _make_ref_id(art, run_at):
    """Generate REF-YYYYMMDD-NNNN style reference ID."""
    date_str = run_at[:10].replace('-', '')
    rank = art.get('stage4_rank') or 0
    # Use source+title hash for non-selected articles
    if not rank:
        import hashlib
        h = hashlib.md5(
            (art.get('source', '') + art.get('title', '')).encode()
        ).hexdigest()[:4].upper()
        return f'REF-{date_str}-{h}'
    return f'REF-{date_str}-{rank:04d}'


def _parse_injection(s1b_reason, art):
    """Extract injection result from stage1b_reason."""
    if not art.get('stage1_passed', False):
        return 'SKIP', 'Not evaluated (failed Stage 1)'
    reason = s1b_reason or ''
    if 'injection' in reason.lower() or 'flagged' in reason.lower():
        return 'FAIL', reason[:100]
    if not art.get('stage1b_passed', True):
        return 'FAIL', reason[:100]
    return 'PASS', 'Clean'


def _parse_sanitized(s1b_reason):
    """Check if article was sanitized."""
    reason = s1b_reason or ''
    return 'sanitiz' in reason.lower() or 'stripped' in reason.lower()


def _parse_gate(art):
    """Extract review gate result."""
    ct = art.get('content_type', 'news')
    rg = art.get('review_gate', '')

    if ct == 'news':
        return 'SKIP', 'News article — gate not applied'
    if not rg:
        return 'SKIP', 'Not evaluated'

    rg_lower = rg.lower()
    if 'passed' in rg_lower:
        return 'PASS', rg[:80]
    if 'rejected' in rg_lower:
        return 'REJECTED', rg[:80]
    if 'skipped' in rg_lower:
        return 'SKIP', rg[:80]
    return rg.upper()[:10], rg[:80]


# ============================================================
# Telegram sender
# ============================================================
def send_trace_to_telegram(xlsx_path: str, run_at: str,
                            article_count: int, selected_count: int) -> bool:
    """Send the forensic trace XLSX to admin Telegram."""
    import requests
    import os

    token = os.getenv('TELEGRAM_BOT_TOKEN', '')
    admin_id = os.getenv('TELEGRAM_ADMIN_ID', '')

    if not token or not admin_id or not xlsx_path:
        return False

    try:
        caption = (
            f'GNI Article Forensic Trace\n'
            f'{run_at[:10]} {run_at[11:16]} UTC\n\n'
            f'Total collected: {article_count}\n'
            f'Selected for AI: {selected_count}\n\n'
            f'Sheets:\n'
            f'1. Selected Articles (quick view)\n'
            f'2. Full Forensic Trace (all {article_count} articles)\n'
            f'3. Rejection Analysis (why each stage rejected)\n'
            f'4. Source Performance (which sources perform best)\n\n'
            f'Color coding: Green=Selected | Yellow=Passed not selected | '
            f'Orange=Rejected 1b | Red=Rejected Stage 1'
        )

        with open(xlsx_path, 'rb') as f:
            resp = requests.post(
                f'https://api.telegram.org/bot{token}/sendDocument',
                data={'chat_id': admin_id, 'caption': caption[:1024]},
                files={'document': (os.path.basename(xlsx_path), f,
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=30
            )

        if resp.status_code == 200:
            print('  OK Forensic trace sent to admin Telegram')
            return True
        else:
            print(f'  Warning: Telegram send failed: {resp.status_code}')
            return False

    except Exception as e:
        print(f'  Warning: Trace Telegram error: {str(e)[:80]}')
        return False
    finally:
        try:
            if xlsx_path and os.path.exists(xlsx_path):
                os.remove(xlsx_path)
        except Exception:
            pass


def run_forensic_trace_pipeline(trace: list, run_id: str,
                                 run_at: str, pipeline_meta: dict) -> bool:
    """
    Main entry point. Generate XLSX and send to Telegram.
    trace: the full article trace list from run_funnel()
    Silent failure -- never breaks the pipeline.
    """
    try:
        print('\n  Generating Article Forensic Trace (XLSX)...')
        selected_count = sum(1 for a in trace if a.get('stage4_selected'))

        xlsx_path = generate_forensic_trace(
            articles=trace,
            run_id=run_id,
            run_at=run_at,
            pipeline_meta=pipeline_meta
        )
        if not xlsx_path:
            return False

        return send_trace_to_telegram(
            xlsx_path=xlsx_path,
            run_at=run_at,
            article_count=len(trace),
            selected_count=selected_count
        )
    except Exception as e:
        print(f'  Warning: Forensic trace pipeline error: {str(e)[:80]}')
        return False
